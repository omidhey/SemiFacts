from itertools import islice

import numpy as np
import pandas as pd
from openpyxl import load_workbook



def xl2pddf(address, sheet_number=1):
    '''
    Convert an Excel file with header to pandas DataFrame
    '''

    #   load Excel file
    wb = load_workbook(address)
    sheet1 = wb[wb.sheetnames[sheet_number]]

    data = sheet1.values
    cols = next(data)[1:]
    data = list(data)
    idx = [r[0] for r in data]
    data = (islice(r, 1, None) for r in data)
    df = pd.DataFrame(data, index=idx, columns=cols)
    return df